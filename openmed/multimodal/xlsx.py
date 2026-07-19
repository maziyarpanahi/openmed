"""Cell-level XLSX redaction with PHI-safe coordinate reporting.

The adapter preserves workbook structure by editing only string-valued data
cells in a normal ``openpyxl`` workbook. Column classification is shared with
the CSV/TSV adapter, while cells in otherwise safe columns still receive
free-text de-identification so embedded PHI is not missed.
"""

from __future__ import annotations

import os
import re
import tempfile
from collections.abc import Callable, Iterable, Mapping, Sequence
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openmed.core.labels import OTHER, normalize_label

from .base import ExtractedDocument
from .exceptions import MissingDependencyError
from .tabular_csv import (
    ACTION_DATE_SHIFT,
    ACTION_FREE_TEXT_REDACT,
    ACTION_KEEP,
    ColumnDecision,
    _date_shift_for_row,
    _merge_policy_options,
    _redact_cell,
    classify_columns,
)

_XLSX_HINT = 'Install with: pip install "openmed[multimodal]".'
_MASK_LABEL_RE = re.compile(r"\[([A-Za-z][A-Za-z0-9_]*)\]")

CellDeidentifier = Callable[[str], Any]


@dataclass(frozen=True)
class XlsxCellRedaction:
    """PHI-safe location and labels for one changed workbook cell."""

    sheet_index: int
    coordinate: str
    labels: tuple[str, ...]

    def to_dict(self) -> dict[str, Any]:
        """Return a report entry that never includes cell or sheet text."""
        return {
            "sheet_index": self.sheet_index,
            "coordinate": self.coordinate,
            "labels": list(self.labels),
        }


@dataclass(frozen=True)
class XlsxRedactionResult:
    """Saved workbook location and PHI-safe cell-level redaction report."""

    output_path: Path
    sheet_count: int
    redactions: tuple[XlsxCellRedaction, ...]

    @property
    def redaction_report(self) -> tuple[dict[str, Any], ...]:
        """Return coordinate/label records without raw cell or sheet text."""
        return tuple(redaction.to_dict() for redaction in self.redactions)

    def to_document(self) -> ExtractedDocument:
        """Bridge the binary workbook result into the multimodal contract."""
        return ExtractedDocument(
            text="",
            metadata={
                "format": "xlsx",
                "sheet_count": self.sheet_count,
                "redaction_count": len(self.redactions),
                "redaction_report": list(self.redaction_report),
            },
        )


def redact_xlsx(
    source: str | os.PathLike[str],
    output_path: str | os.PathLike[str],
    *,
    policy: Any | None = None,
    models: Any | None = None,
    lang: str = "en",
    header_row: int = 1,
    header_heuristics: Mapping[str, str] | None = None,
    action_overrides: Mapping[str, str] | None = None,
    date_shift_days: int | None = None,
    date_shift_seed: str = "openmed-xlsx-v1",
    keep_year: bool = True,
    sample_size: int = 50,
    cell_deidentifier: CellDeidentifier | None = None,
) -> XlsxRedactionResult:
    """Redact PHI in XLSX string cells while preserving workbook structure.

    The first row is treated as the column-header row by default. Its names and
    sampled string values are classified with the CSV/TSV PHI-column rules.
    Formula cells and non-string values are never passed to de-identification
    and are never changed. String cells not covered by a PHI-column action are
    still analyzed individually for free-text PHI.

    Args:
        source: Input ``.xlsx`` workbook path.
        output_path: Destination ``.xlsx`` path. It must differ from ``source``.
        policy: Optional tabular mapping with ``header_heuristics`` and
            ``action_overrides``, or a core policy profile name for the default
            cell de-identifier.
        models: Optional callable, mapping, or object exposing
            ``cell_deidentifier``, ``deidentifier``, or ``text_redactor``.
        lang: Language passed to the default OpenMed de-identification path.
        header_row: One-based row containing column names on every worksheet.
        header_heuristics: Additional CSV-compatible header label mappings.
        action_overrides: Additional CSV-compatible column action mappings.
        date_shift_days: Optional fixed shift for string date columns.
        date_shift_seed: Seed for deterministic per-row date shifts.
        keep_year: Preserve years when shifting string dates.
        sample_size: Non-empty string cells sampled per column.
        cell_deidentifier: Optional callable for free-text cells. It may return
            an OpenMed ``DeidentificationResult``, ``(text, labels)``, a mapping
            with ``deidentified_text`` and labels/entities, or a redacted string.

    Returns:
        The saved output path and a report containing only sheet indexes, cell
        coordinates, and canonical labels.

    Raises:
        ValueError: If paths or row/sample options are invalid.
        MissingDependencyError: If ``openpyxl`` is not installed.
    """
    source_path = Path(source)
    destination = Path(output_path)
    _validate_paths(source_path, destination)
    if header_row < 1:
        raise ValueError("header_row must be at least 1")
    if sample_size < 1:
        raise ValueError("sample_size must be at least 1")

    load_workbook = _load_workbook_function()
    workbook = load_workbook(
        filename=source_path,
        read_only=False,
        data_only=False,
        keep_links=True,
    )
    resolved_deidentifier = _resolve_cell_deidentifier(
        explicit=cell_deidentifier,
        models=models,
        policy=policy,
        lang=lang,
    )
    merged_headers, merged_actions = _merge_policy_options(
        policy,
        header_heuristics=header_heuristics,
        action_overrides=action_overrides,
    )
    sheet_count = len(workbook.worksheets)
    redactions: list[XlsxCellRedaction] = []

    try:
        for sheet_index, worksheet in enumerate(workbook.worksheets):
            decisions = _worksheet_decisions(
                worksheet,
                header_row=header_row,
                header_heuristics=merged_headers,
                action_overrides=merged_actions,
                sample_size=sample_size,
            )
            _redact_worksheet(
                worksheet,
                sheet_index=sheet_index,
                header_row=header_row,
                decisions=decisions,
                deidentifier=resolved_deidentifier,
                date_shift_days=date_shift_days,
                date_shift_seed=date_shift_seed,
                keep_year=keep_year,
                lang=lang,
                redactions=redactions,
            )
        _save_workbook_atomically(workbook, destination)
    finally:
        workbook.close()

    return XlsxRedactionResult(
        output_path=destination,
        sheet_count=sheet_count,
        redactions=tuple(redactions),
    )


def _validate_paths(source: Path, destination: Path) -> None:
    if source.suffix.lower() != ".xlsx":
        raise ValueError("source must be an .xlsx workbook")
    if destination.suffix.lower() != ".xlsx":
        raise ValueError("output_path must end in .xlsx")
    if source.resolve() == destination.resolve():
        raise ValueError("output_path must differ from source")
    if not source.is_file():
        raise FileNotFoundError(source)
    if not destination.parent.is_dir():
        raise FileNotFoundError(destination.parent)


def _load_workbook_function() -> Callable[..., Any]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise MissingDependencyError(
            dependency="openpyxl",
            instruction=_XLSX_HINT,
        ) from exc
    return load_workbook


def _worksheet_decisions(
    worksheet: Any,
    *,
    header_row: int,
    header_heuristics: Mapping[str, str] | None,
    action_overrides: Mapping[str, str] | None,
    sample_size: int,
) -> tuple[ColumnDecision, ...]:
    if header_row > worksheet.max_row or _worksheet_is_empty(worksheet):
        return ()

    headers = tuple(
        _header_name(worksheet.cell(row=header_row, column=column).value, column)
        for column in range(1, worksheet.max_column + 1)
    )
    rows = _sample_string_rows(
        worksheet,
        header_row=header_row,
        width=worksheet.max_column,
        sample_size=sample_size,
    )
    return classify_columns(
        headers,
        rows,
        header_heuristics=header_heuristics,
        action_overrides=action_overrides,
        sample_size=sample_size,
    )


def _sample_string_rows(
    worksheet: Any,
    *,
    header_row: int,
    width: int,
    sample_size: int,
) -> tuple[tuple[str, ...], ...]:
    counts = [0] * width
    rows: list[tuple[str, ...]] = []
    for row_index in range(header_row + 1, worksheet.max_row + 1):
        values = tuple(
            _classification_value(worksheet.cell(row=row_index, column=column))
            for column in range(1, width + 1)
        )
        contributes = tuple(
            bool(value.strip()) and counts[index] < sample_size
            for index, value in enumerate(values)
        )
        if not any(contributes):
            continue
        rows.append(values)
        for index, contributes_to_column in enumerate(contributes):
            if contributes_to_column:
                counts[index] += 1
        if all(count >= sample_size for count in counts):
            break
    return tuple(rows)


def _worksheet_is_empty(worksheet: Any) -> bool:
    return (
        worksheet.max_row == 1
        and worksheet.max_column == 1
        and worksheet.cell(row=1, column=1).value is None
    )


def _header_name(value: Any, column: int) -> str:
    if isinstance(value, str) and value.strip():
        return value.strip()
    return f"column_{column}"


def _classification_value(cell: Any) -> str:
    if cell.data_type == "f" or not isinstance(cell.value, str):
        return ""
    return cell.value


def _redact_worksheet(
    worksheet: Any,
    *,
    sheet_index: int,
    header_row: int,
    decisions: Sequence[ColumnDecision],
    deidentifier: CellDeidentifier,
    date_shift_days: int | None,
    date_shift_seed: str,
    keep_year: bool,
    lang: str,
    redactions: list[XlsxCellRedaction],
) -> None:
    for row_index in range(header_row + 1, worksheet.max_row + 1):
        row_values = tuple(
            _classification_value(worksheet.cell(row=row_index, column=column))
            for column in range(1, len(decisions) + 1)
        )
        row_shift_days: int | None = None
        for decision in decisions:
            cell = worksheet.cell(row=row_index, column=decision.index + 1)
            if cell.data_type == "f" or not isinstance(cell.value, str):
                continue

            original = cell.value
            if decision.action in {ACTION_KEEP, ACTION_FREE_TEXT_REDACT}:
                redacted, labels = _run_cell_deidentifier(original, deidentifier)
            else:
                if decision.action == ACTION_DATE_SHIFT and row_shift_days is None:
                    row_shift_days = _date_shift_for_row(
                        row_values,
                        row_index=row_index - header_row - 1,
                        fixed_days=date_shift_days,
                        seed=date_shift_seed,
                    )
                redacted, changed = _redact_cell(
                    original,
                    decision,
                    date_shift_days=row_shift_days,
                    keep_year=keep_year,
                    lang=lang,
                    text_redactor=None,
                )
                labels = (
                    (_normalize_report_label(decision.canonical_label or "PHI"),)
                    if changed
                    else ()
                )

            if redacted == original:
                continue
            cell.value = redacted
            redactions.append(
                XlsxCellRedaction(
                    sheet_index=sheet_index,
                    coordinate=cell.coordinate,
                    labels=_unique_labels(labels) or ("PHI",),
                )
            )


def _resolve_cell_deidentifier(
    *,
    explicit: CellDeidentifier | None,
    models: Any | None,
    policy: Any | None,
    lang: str,
) -> CellDeidentifier:
    if explicit is not None:
        return explicit
    if callable(models):
        return models
    if isinstance(models, Mapping):
        for key in ("cell_deidentifier", "deidentifier", "text_redactor"):
            candidate = models.get(key)
            if callable(candidate):
                return candidate
    for attribute in ("cell_deidentifier", "deidentifier", "text_redactor"):
        candidate = getattr(models, attribute, None)
        if callable(candidate):
            return candidate

    core_policy = policy if isinstance(policy, str) else None

    def default_deidentifier(text: str) -> Any:
        from openmed.core.pii import deidentify

        return deidentify(text, method="mask", lang=lang, policy=core_policy)

    return default_deidentifier


def _run_cell_deidentifier(
    value: str,
    deidentifier: CellDeidentifier,
) -> tuple[str, tuple[str, ...]]:
    result = deidentifier(value)
    redacted, labels = _coerce_deidentification_result(result)
    if redacted == value:
        return value, ()
    if not labels:
        labels = tuple(_MASK_LABEL_RE.findall(redacted))
    return redacted, _unique_labels(labels) or ("PHI",)


def _coerce_deidentification_result(result: Any) -> tuple[str, tuple[str, ...]]:
    if isinstance(result, str):
        return result, ()
    if isinstance(result, tuple) and len(result) == 2:
        return str(result[0]), _coerce_labels(result[1])
    if isinstance(result, Mapping):
        redacted = result.get("deidentified_text", result.get("text"))
        if redacted is None:
            raise TypeError("cell de-identifier mapping must include deidentified_text")
        labels = _coerce_labels(result.get("labels"))
        if not labels:
            labels = _labels_from_entities(result.get("pii_entities", ()))
        return str(redacted), labels

    redacted = getattr(result, "deidentified_text", None)
    if redacted is None:
        raise TypeError(
            "cell de-identifier must return text, (text, labels), a mapping, "
            "or an object with deidentified_text"
        )
    labels = _labels_from_entities(getattr(result, "pii_entities", ()))
    return str(redacted), labels


def _labels_from_entities(entities: Iterable[Any] | None) -> tuple[str, ...]:
    if entities is None:
        return ()
    labels: list[str] = []
    for entity in entities:
        if isinstance(entity, Mapping):
            label = (
                entity.get("canonical_label")
                or entity.get("label")
                or entity.get("entity_type")
            )
        else:
            label = (
                getattr(entity, "canonical_label", None)
                or getattr(entity, "label", None)
                or getattr(entity, "entity_type", None)
            )
        if label:
            labels.append(str(label))
    return tuple(labels)


def _coerce_labels(labels: Any) -> tuple[str, ...]:
    if labels is None:
        return ()
    if isinstance(labels, str):
        return (labels,)
    try:
        return tuple(str(label) for label in labels)
    except TypeError:
        return (str(labels),)


def _normalize_report_label(label: str) -> str:
    canonical = normalize_label(str(label))
    if canonical != OTHER:
        return canonical
    return "PHI" if str(label).strip().upper() == "PHI" else OTHER


def _unique_labels(labels: Iterable[str]) -> tuple[str, ...]:
    return tuple(
        dict.fromkeys(
            _normalize_report_label(label) for label in labels if str(label).strip()
        )
    )


def _save_workbook_atomically(workbook: Any, destination: Path) -> None:
    temporary_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            prefix=f".{destination.stem}.",
            suffix=".xlsx",
            dir=destination.parent,
            delete=False,
        ) as temporary:
            temporary_path = Path(temporary.name)
        workbook.save(temporary_path)
        os.replace(temporary_path, destination)
        temporary_path = None
    finally:
        if temporary_path is not None:
            temporary_path.unlink(missing_ok=True)
